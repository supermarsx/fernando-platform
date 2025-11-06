import csv
import io
import json
import zipfile
from datetime import datetime
from typing import Dict, List, Any, Optional, Union
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, func
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd

from app.models.enterprise import ExportJob
from app.models.job import Job
from app.models.document import Document
from app.models.extraction import ExtractionRun, ExtractionField
from app.services.enterprise_service import EnterpriseService


class ExportImportService:
    """Service for handling data export/import in multiple formats"""
    
    def __init__(self, db: Session):
        self.db = db
        self.enterprise_service = EnterpriseService(db)
    
    def create_export(self, tenant_id: str, user_id: str, export_type: str,
                     export_format: str, job_id: str = None, 
                     filters: Dict[str, Any] = None) -> ExportJob:
        """Create a new export job"""
        export_job = self.enterprise_service.create_export_job(
            tenant_id=tenant_id,
            user_id=user_id,
            export_type=export_type,
            export_format=export_format,
            job_id=job_id,
            filters=filters or {}
        )
        return export_job
    
    def execute_export(self, export_job_id: str) -> Dict[str, Any]:
        """Execute an export job and return file data"""
        export_job = self.db.query(ExportJob).filter(
            ExportJob.export_id == export_job_id
        ).first()
        
        if not export_job:
            return {"error": "Export job not found"}
        
        try:
            # Update status to processing
            export_job.status = "processing"
            self.db.commit()
            
            # Get data based on export type
            data = self._get_export_data(export_job)
            
            # Generate file based on format
            file_data, filename, mime_type = self._generate_export_file(
                data, export_job.export_type, export_job.export_format
            )
            
            # Update export job
            export_job.status = "completed"
            export_job.file_size_bytes = len(file_data)
            export_job.records_count = len(data) if isinstance(data, list) else 1
            export_job.completed_at = datetime.utcnow()
            self.db.commit()
            
            return {
                "success": True,
                "file_data": file_data,
                "filename": filename,
                "mime_type": mime_type,
                "export_job": export_job
            }
            
        except Exception as e:
            export_job.status = "failed"
            export_job.error_message = str(e)
            self.db.commit()
            
            return {"error": str(e)}
    
    def _get_export_data(self, export_job: ExportJob) -> Union[List[Dict], Dict]:
        """Get data for export based on type and filters"""
        filters = export_job.filters or {}
        
        if export_job.export_type == "jobs":
            return self._get_jobs_data(export_job.tenant_id, filters)
        elif export_job.export_type == "documents":
            return self._get_documents_data(export_job.tenant_id, filters)
        elif export_job.export_type == "extractions":
            return self._get_extractions_data(export_job.tenant_id, filters)
        elif export_job.export_type == "audit_logs":
            return self._get_audit_data(export_job.tenant_id, filters)
        elif export_job.export_type == "tenant_summary":
            return self._get_tenant_summary(export_job.tenant_id, filters)
        else:
            return []
    
    def _get_jobs_data(self, tenant_id: str, filters: Dict[str, Any]) -> List[Dict]:
        """Get jobs data for export"""
        query = self.db.query(Job).filter(Job.tenant_id == tenant_id)
        
        # Apply filters
        if filters.get("status"):
            query = query.filter(Job.status.in_(filters["status"]))
        if filters.get("date_from"):
            query = query.filter(Job.created_at >= filters["date_from"])
        if filters.get("date_to"):
            query = query.filter(Job.created_at <= filters["date_to"])
        
        jobs = query.order_by(Job.created_at.desc()).all()
        
        data = []
        for job in jobs:
            data.append({
                "job_id": job.job_id,
                "status": job.status,
                "priority": job.priority,
                "queue_name": job.queue_name,
                "uploaded_by": job.uploaded_by,
                "assigned_to": job.assigned_to,
                "created_at": job.created_at.isoformat(),
                "started_at": job.started_at.isoformat() if job.started_at else None,
                "finished_at": job.finished_at.isoformat() if job.finished_at else None,
                "estimated_duration": job.estimated_duration,
                "actual_duration": job.actual_duration,
                "retry_count": job.retry_count,
                "error_code": job.error_code,
                "progress_percentage": job.progress_percentage,
                "metadata": json.dumps(job.metadata or {})
            })
        
        return data
    
    def _get_documents_data(self, tenant_id: str, filters: Dict[str, Any]) -> List[Dict]:
        """Get documents data for export"""
        query = self.db.query(Document).filter(Document.tenant_id == tenant_id)
        
        # Join with jobs to get more details
        query = query.join(Job, Document.job_id == Job.job_id)
        
        # Apply filters
        if filters.get("mime_type"):
            query = query.filter(Document.mime_type.in_(filters["mime_type"]))
        if filters.get("date_from"):
            query = query.filter(Document.created_at >= filters["date_from"])
        if filters.get("date_to"):
            query = query.filter(Document.created_at <= filters["date_to"])
        
        documents = query.order_by(Document.created_at.desc()).all()
        
        data = []
        for doc in documents:
            data.append({
                "document_id": doc.document_id,
                "job_id": doc.job_id,
                "original_filename": doc.original_filename,
                "mime_type": doc.mime_type,
                "pages": doc.pages,
                "file_size_bytes": doc.file_size_bytes,
                "checksum_sha256": doc.checksum_sha256,
                "uploaded_by": doc.uploaded_by,
                "is_confidential": doc.is_confidential,
                "retention_period_days": doc.retention_period_days,
                "tags": ", ".join(doc.tags or []),
                "created_at": doc.created_at.isoformat(),
                "processed_at": doc.processed_at.isoformat() if doc.processed_at else None,
                "archived_at": doc.archived_at.isoformat() if doc.archived_at else None,
                "storage_url": doc.storage_url
            })
        
        return data
    
    def _get_extractions_data(self, tenant_id: str, filters: Dict[str, Any]) -> List[Dict]:
        """Get extractions data for export"""
        query = self.db.query(ExtractionRun).join(Job).filter(Job.tenant_id == tenant_id)
        
        # Apply filters
        if filters.get("status"):
            query = query.filter(ExtractionRun.status.in_(filters["status"]))
        if filters.get("date_from"):
            query = query.filter(ExtractionRun.created_at >= filters["date_from"])
        if filters.get("date_to"):
            query = query.filter(ExtractionRun.created_at <= filters["date_to"])
        
        runs = query.order_by(ExtractionRun.created_at.desc()).all()
        
        data = []
        for run in runs:
            data.append({
                "extraction_id": run.extraction_id,
                "job_id": run.job_id,
                "document_id": run.document_id,
                "status": run.status,
                "ocr_text_length": len(run.ocr_text) if run.ocr_text else 0,
                "created_at": run.created_at.isoformat(),
                "completed_at": run.completed_at.isoformat() if run.completed_at else None,
                "confidence_score": run.confidence_score
            })
        
        return data
    
    def _get_audit_data(self, tenant_id: str, filters: Dict[str, Any]) -> List[Dict]:
        """Get audit data for export"""
        from app.models.enterprise import AuditTrail
        
        query = self.db.query(AuditTrail).filter(AuditTrail.tenant_id == tenant_id)
        
        # Apply filters
        if filters.get("user_id"):
            query = query.filter(AuditTrail.user_id == filters["user_id"])
        if filters.get("action"):
            query = query.filter(AuditTrail.action.in_(filters["action"]))
        if filters.get("risk_level"):
            query = query.filter(AuditTrail.risk_level.in_(filters["risk_level"]))
        if filters.get("date_from"):
            query = query.filter(AuditTrail.created_at >= filters["date_from"])
        if filters.get("date_to"):
            query = query.filter(AuditTrail.created_at <= filters["date_to"])
        
        audit_logs = query.order_by(AuditTrail.created_at.desc()).all()
        
        data = []
        for log in audit_logs:
            data.append({
                "audit_id": log.audit_id,
                "user_id": log.user_id,
                "action": log.action,
                "resource_type": log.resource_type,
                "resource_id": log.resource_id,
                "ip_address": log.ip_address,
                "user_agent": log.user_agent,
                "risk_level": log.risk_level,
                "compliance_tags": ", ".join(log.compliance_tags or []),
                "created_at": log.created_at.isoformat(),
                "old_values": json.dumps(log.old_values) if log.old_values else None,
                "new_values": json.dumps(log.new_values) if log.new_values else None
            })
        
        return data
    
    def _get_tenant_summary(self, tenant_id: str, filters: Dict[str, Any]) -> Dict[str, Any]:
        """Get tenant summary for export"""
        # Get various counts and statistics
        job_stats = self.db.query(
            Job.status,
            func.count(Job.job_id)
        ).filter(Job.tenant_id == tenant_id).group_by(Job.status).all()
        
        doc_stats = self.db.query(
            Document.mime_type,
            func.count(Document.document_id),
            func.sum(Document.file_size_bytes)
        ).filter(Document.tenant_id == tenant_id).group_by(Document.mime_type).all()
        
        user_count = self.db.query(func.count(Job.uploaded_by.distinct())).filter(
            Job.tenant_id == tenant_id
        ).scalar()
        
        return {
            "tenant_id": tenant_id,
            "export_date": datetime.utcnow().isoformat(),
            "job_statistics": dict(job_stats),
            "document_statistics": [
                {
                    "mime_type": mime_type,
                    "count": count,
                    "total_size_bytes": total_size or 0
                }
                for mime_type, count, total_size in doc_stats
            ],
            "unique_users": user_count,
            "period": filters.get("period", "all_time")
        }
    
    def _generate_export_file(self, data: Union[List[Dict], Dict], 
                            export_type: str, export_format: str) -> tuple:
        """Generate export file in specified format"""
        if export_format == "csv":
            return self._generate_csv(data, export_type)
        elif export_format == "excel":
            return self._generate_excel(data, export_type)
        elif export_format == "json":
            return self._generate_json(data)
        elif export_format == "pdf":
            return self._generate_pdf_report(data, export_type)
        else:
            raise ValueError(f"Unsupported export format: {export_format}")
    
    def _generate_csv(self, data: List[Dict], export_type: str) -> tuple:
        """Generate CSV file"""
        if not data:
            csv_content = ""
        else:
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
            csv_content = output.getvalue()
            output.close()
        
        filename = f"{export_type}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
        return csv_content.encode('utf-8'), filename, "text/csv"
    
    def _generate_excel(self, data: List[Dict], export_type: str) -> tuple:
        """Generate Excel file with formatting"""
        wb = Workbook()
        ws = wb.active
        ws.title = export_type.title()
        
        if not data:
            ws.append(["No data available"])
        else:
            # Add headers
            headers = list(data[0].keys())
            ws.append(headers)
            
            # Style headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Add data
            for row in data:
                ws.append(list(row.values()))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        filename = f"{export_type}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return excel_buffer.read(), filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    def _generate_json(self, data: Union[List[Dict], Dict]) -> tuple:
        """Generate JSON file"""
        json_content = json.dumps(data, indent=2, default=str)
        filename = f"export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json"
        return json_content.encode('utf-8'), filename, "application/json"
    
    def _generate_pdf_report(self, data: Union[List[Dict], Dict], export_type: str) -> tuple:
        """Generate PDF report (placeholder - would need reportlab or similar)"""
        # For now, return a simple text file as PDF placeholder
        # In a real implementation, you'd use reportlab or similar
        content = f"PDF Report - {export_type}\nGenerated: {datetime.utcnow()}\n\n"
        
        if isinstance(data, dict):
            content += json.dumps(data, indent=2, default=str)
        else:
            content += f"Total records: {len(data)}\n"
            if data:
                content += "Sample record:\n"
                content += json.dumps(data[0], indent=2, default=str)
        
        filename = f"{export_type}_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.txt"
        return content.encode('utf-8'), filename, "text/plain"
    
    def import_data(self, tenant_id: str, user_id: str, file_data: bytes, 
                   filename: str, import_type: str) -> Dict[str, Any]:
        """Import data from file"""
        try:
            if filename.endswith('.csv'):
                return self._import_csv(tenant_id, user_id, file_data, import_type)
            elif filename.endswith('.json'):
                return self._import_json(tenant_id, user_id, file_data, import_type)
            elif filename.endswith(('.xlsx', '.xls')):
                return self._import_excel(tenant_id, user_id, file_data, import_type)
            else:
                return {"error": "Unsupported file format"}
        except Exception as e:
            return {"error": f"Import failed: {str(e)}"}
    
    def _import_csv(self, tenant_id: str, user_id: str, file_data: bytes, 
                   import_type: str) -> Dict[str, Any]:
        """Import data from CSV file"""
        csv_content = file_data.decode('utf-8')
        csv_reader = csv.DictReader(io.StringIO(csv_content))
        data = list(csv_reader)
        
        if import_type == "jobs":
            return self._process_job_import(tenant_id, user_id, data)
        elif import_type == "users":
            return self._process_user_import(tenant_id, user_id, data)
        else:
            return {"error": f"Unsupported import type: {import_type}"}
    
    def _import_json(self, tenant_id: str, user_id: str, file_data: bytes,
                    import_type: str) -> Dict[str, Any]:
        """Import data from JSON file"""
        json_content = file_data.decode('utf-8')
        data = json.loads(json_content)
        
        if import_type == "jobs":
            return self._process_job_import(tenant_id, user_id, data)
        elif import_type == "users":
            return self._process_user_import(tenant_id, user_id, data)
        else:
            return {"error": f"Unsupported import type: {import_type}"}
    
    def _import_excel(self, tenant_id: str, user_id: str, file_data: bytes,
                     import_type: str) -> Dict[str, Any]:
        """Import data from Excel file"""
        excel_buffer = io.BytesIO(file_data)
        df = pd.read_excel(excel_buffer)
        data = df.to_dict('records')
        
        if import_type == "jobs":
            return self._process_job_import(tenant_id, user_id, data)
        elif import_type == "users":
            return self._process_user_import(tenant_id, user_id, data)
        else:
            return {"error": f"Unsupported import type: {import_type}"}
    
    def _process_job_import(self, tenant_id: str, user_id: str, data: List[Dict]) -> Dict[str, Any]:
        """Process job import data"""
        imported_count = 0
        errors = []
        
        for row in data:
            try:
                # Validate required fields
                if not row.get('original_filename'):
                    errors.append(f"Missing original_filename in row")
                    continue
                
                # Create job record (simplified)
                job = Job(
                    tenant_id=tenant_id,
                    status="queued",
                    uploaded_by=user_id,
                    priority=int(row.get('priority', 0)),
                    queue_name=row.get('queue_name', 'default'),
                    metadata=row
                )
                self.db.add(job)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Error processing row: {str(e)}")
        
        self.db.commit()
        
        return {
            "success": True,
            "imported_count": imported_count,
            "errors": errors
        }
    
    def _process_user_import(self, tenant_id: str, user_id: str, data: List[Dict]) -> Dict[str, Any]:
        """Process user import data"""
        from app.models.enterprise import UserEnterprise
        
        imported_count = 0
        errors = []
        
        for row in data:
            try:
                # Validate required fields
                if not row.get('email') or not row.get('full_name'):
                    errors.append(f"Missing required fields (email, full_name) in row")
                    continue
                
                # Check if user already exists
                existing_user = self.db.query(UserEnterprise).filter(
                    and_(
                        UserEnterprise.tenant_id == tenant_id,
                        UserEnterprise.email == row['email']
                    )
                ).first()
                
                if existing_user:
                    errors.append(f"User with email {row['email']} already exists")
                    continue
                
                # Create user record
                user = UserEnterprise(
                    tenant_id=tenant_id,
                    email=row['email'],
                    full_name=row['full_name'],
                    password_hash="",  # Would need proper password hashing
                    status=row.get('status', 'active')
                )
                self.db.add(user)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Error processing row: {str(e)}")
        
        self.db.commit()
        
        return {
            "success": True,
            "imported_count": imported_count,
            "errors": errors
        }
